from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

OUTPUT_DIR = Path("output")


GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE",
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE",
)


COLUMN_MAPPING = {
    "roe_min": "return_on_equity_pct",
    "free_cash_flow_min": "free_cash_flow_cr",
    "revenue_cagr_5yr_min": "revenue_cagr_5yr",
    "pat_cagr_5yr_min": "pat_cagr_5yr",
    "operating_profit_margin_min": "operating_profit_margin_pct",
    "interest_coverage_min": "interest_coverage",
    "net_profit_min": "net_profit",
    "eps_cagr_5yr_min": "eps_cagr_5yr",
    "asset_turnover_min": "asset_turnover",
    "sales_min": "sales",
    "debt_to_equity_max": "debt_to_equity",
    "dividend_payout_max": "dividend_payout",
}


def export_all_screeners(
    results: dict[str, pd.DataFrame],
    presets: dict,
) -> Path:

    OUTPUT_DIR.mkdir(exist_ok=True)

    output_path = OUTPUT_DIR / "screener_output.xlsx"

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:

        for sheet_name, df in results.items():

            df.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

            worksheet = writer.sheets[sheet_name[:31]]

            preset = presets[sheet_name]

            headers = {cell.value: cell.column for cell in worksheet[1]}

            for filter_name, threshold in preset.items():

                if threshold is None:
                    continue

                if filter_name not in COLUMN_MAPPING:
                    continue

                column_name = COLUMN_MAPPING[filter_name]

                if column_name not in headers:
                    continue

                col = headers[column_name]

                is_max_filter = filter_name.endswith("_max")

                for row in range(
                    2,
                    worksheet.max_row + 1,
                ):

                    cell = worksheet.cell(
                        row=row,
                        column=col,
                    )

                    value = cell.value

                    if value is None:
                        continue

                    if is_max_filter:

                        if value <= threshold:
                            cell.fill = GREEN_FILL
                        else:
                            cell.fill = RED_FILL

                    else:

                        if value >= threshold:
                            cell.fill = GREEN_FILL
                        else:
                            cell.fill = RED_FILL

    return output_path
