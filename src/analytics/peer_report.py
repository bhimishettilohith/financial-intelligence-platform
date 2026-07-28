from __future__ import annotations

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

DB_PATH = "data/nifty100.db"

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

OUTPUT_FILE = OUTPUT_DIR / "peer_comparison.xlsx"


METRIC_LABELS = {
    "return_on_equity_pct": "ROE",
    "return_on_capital_employed_pct": "ROCE",
    "net_profit_margin_pct": "Net Profit Margin",
    "debt_to_equity": "Debt/Equity",
    "free_cash_flow_cr": "Free Cash Flow",
    "pat_cagr_5yr": "PAT CAGR (5Y)",
    "revenue_cagr_5yr": "Revenue CAGR (5Y)",
    "eps_cagr_5yr": "EPS CAGR (5Y)",
    "interest_coverage": "Interest Coverage",
    "asset_turnover": "Asset Turnover",
}

METRIC_ORDER = list(METRIC_LABELS.keys())


GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAD3",
)

BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD966",
)

HEADER_FONT = Font(
    bold=True,
)

MEDIAN_FONT = Font(
    bold=True,
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class PeerComparisonReport:

    def __init__(self):

        self.conn = sqlite3.connect(DB_PATH)

        self.workbook = Workbook()

        default = self.workbook.active
        self.workbook.remove(default)

        self.load_data()

    def load_data(self):

        self.peer_groups = pd.read_sql(
            """
            SELECT
                company_id,
                peer_group_name,
                is_benchmark
            FROM peer_groups
            """,
            self.conn,
        )

        self.company_master = pd.read_sql(
            """
            SELECT
                id,
                company_name
            FROM companies
            """,
            self.conn,
        )

        self.percentiles = pd.read_sql(
            """
            SELECT
                company_id,
                peer_group_name,
                metric,
                value,
                percentile_rank
            FROM peer_percentiles
            """,
            self.conn,
        )

        self.company_master.rename(
            columns={"id": "company_id"},
            inplace=True,
        )

    def build_peer_dataframe(
        self,
        peer_group: str,
    ) -> pd.DataFrame:

        df = self.percentiles[self.percentiles["peer_group_name"] == peer_group].copy()

        value_table = df.pivot_table(
            index="company_id",
            columns="metric",
            values="value",
        ).reset_index()

        percentile_table = df.pivot_table(
            index="company_id",
            columns="metric",
            values="percentile_rank",
        ).reset_index()

        percentile_table.rename(
            columns={metric: f"{metric}_pct" for metric in METRIC_ORDER},
            inplace=True,
        )

        result = value_table.merge(
            percentile_table,
            on="company_id",
        )

        result = result.merge(
            self.company_master,
            on="company_id",
            how="left",
        )

        result = result.merge(
            self.peer_groups,
            on="company_id",
            how="left",
        )

        return result

    def create_sheet(
        self,
        peer_group: str,
    ):

        ws = self.workbook.create_sheet(title=peer_group[:31])

        headers = [
            "Company ID",
            "Company Name",
        ]

        for metric in METRIC_ORDER:

            headers.append(METRIC_LABELS[metric])

            headers.append(f"{METRIC_LABELS[metric]} Percentile")

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(
                row=1,
                column=col,
                value=header,
            )

            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        df = self.build_peer_dataframe(peer_group)

        row_num = 2

        for _, row in df.iterrows():

            values = [
                row["company_id"],
                row["company_name"],
            ]

            for metric in METRIC_ORDER:

                values.append(row.get(metric))

                pct = row.get(f"{metric}_pct")

                if pd.notna(pct):
                    pct *= 100

                values.append(pct)

            for col_num, value in enumerate(values, start=1):

                cell = ws.cell(
                    row=row_num,
                    column=col_num,
                    value=value,
                )

                cell.border = THIN_BORDER

                if col_num > 2:
                    cell.alignment = Alignment(horizontal="center")

            # ---------------------------------
            # Percentile colour formatting
            # ---------------------------------

            percentile_columns = range(
                4,
                len(headers) + 1,
                2,
            )

            for col in percentile_columns:

                cell = ws.cell(
                    row=row_num,
                    column=col,
                )

                if cell.value is None:
                    continue

                if cell.value >= 75:

                    cell.fill = GREEN_FILL

                elif cell.value <= 25:

                    cell.fill = RED_FILL

                else:

                    cell.fill = YELLOW_FILL

            # ---------------------------------
            # Benchmark company highlighting
            # ---------------------------------

            if row["is_benchmark"] == 1:

                for col in range(
                    1,
                    len(headers) + 1,
                ):

                    ws.cell(
                        row=row_num,
                        column=col,
                    ).fill = BENCHMARK_FILL

            row_num += 1

        return ws, df

    def add_median_row(
        self,
        ws,
        df: pd.DataFrame,
    ):

        row = ws.max_row + 1

        # Style the entire summary row
        for col in range(1, ws.max_column + 1):

            cell = ws.cell(
                row=row,
                column=col,
            )

            cell.fill = HEADER_FILL
            cell.font = MEDIAN_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        ws.cell(
            row=row,
            column=1,
            value="Peer Median",
        )

        metric_col = 3

        for metric in METRIC_ORDER:

            median = df[metric].median()

            ws.cell(
                row=row,
                column=metric_col,
                value=round(median, 2),
            )

            # Leave percentile column blank
            ws.cell(
                row=row,
                column=metric_col + 1,
                value=None,
            )

            metric_col += 2

    @staticmethod
    def autofit_columns(ws):

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )

            ws.column_dimensions[column_cells[0].column_letter].width = min(
                max(length + 2, 12),
                35,
            )

    def generate(self):

        peer_groups = sorted(self.peer_groups["peer_group_name"].dropna().unique())

        for peer_group in peer_groups:

            ws, df = self.create_sheet(peer_group)

            self.add_median_row(
                ws,
                df,
            )

            ws.freeze_panes = "A2"

            self.autofit_columns(ws)

        self.workbook.save(OUTPUT_FILE)

        print("=" * 70)
        print("PEER COMPARISON REPORT")
        print("=" * 70)
        print(f"Workbook saved to: {OUTPUT_FILE}")
        print(f"Worksheets created: {len(peer_groups)}")

    def close(self):

        self.conn.close()


def main():

    report = PeerComparisonReport()

    try:

        report.generate()

    finally:

        report.close()


if __name__ == "__main__":
    main()
